"""
엑셀 서비스.
엑셀 파일에서 작업자 목록을 파싱하여 DB에 일괄 등록한다.
system_id를 자동 발급한다.
"""

import io
import logging
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.user import User

logger = logging.getLogger(__name__)


def generate_system_id(db: Session) -> str:
    """
    새로운 system_id를 자동 발급한다.
    형식: USR-YYYYMMDD-NNNN (오늘 날짜 기준 순번)

    Args:
        db: DB 세션

    Returns:
        발급된 system_id 문자열 (예: USR-20260312-0001)
    """
    today_str = date.today().strftime("%Y%m%d")
    prefix = f"USR-{today_str}-"

    # 오늘 발급된 마지막 순번 조회
    last_user = (
        db.query(User)
        .filter(User.system_id.like(f"{prefix}%"))
        .order_by(User.system_id.desc())
        .first()
    )

    if last_user:
        # 마지막 순번에서 +1
        last_num = int(last_user.system_id.split("-")[-1])
        next_num = last_num + 1
    else:
        next_num = 1

    return f"{prefix}{next_num:04d}"


def parse_and_import_users(db: Session, file_data: bytes, group_id: int) -> dict:
    """
    엑셀 파일을 파싱하여 작업자를 일괄 등록한다.
    system_id는 서버에서 자동 발급한다.

    엑셀 형식 (헤더 포함):
    | 이름 | 사원번호(선택) | 언어(선택) |

    Args:
        db: DB 세션
        file_data: 엑셀 파일 바이너리 데이터
        group_id: 등록할 그룹 ID

    Returns:
        dict: { total, created, skipped, errors }
    """
    wb = load_workbook(filename=io.BytesIO(file_data), read_only=True)
    ws = wb.active

    total = 0
    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        total += 1
        name = str(row[0]).strip()
        emp_no = str(row[1]).strip() if len(row) > 1 and row[1] else None
        language = str(row[2]).strip() if len(row) > 2 and row[2] else "ko"

        if not name:
            errors.append(f"행 {row_idx}: 이름이 비어있습니다")
            skipped += 1
            continue

        # 사원번호 중복 확인
        if emp_no:
            existing = db.query(User).filter(User.emp_no == emp_no).first()
            if existing:
                errors.append(f"행 {row_idx}: 이미 등록된 사원번호 ({emp_no} - {existing.name})")
                skipped += 1
                continue

        try:
            system_id = generate_system_id(db)
            user = User(
                system_id=system_id,
                emp_no=emp_no if emp_no else None,
                name=name,
                language=language,
                group_id=group_id,
            )
            db.add(user)
            db.flush()
            created += 1
        except IntegrityError:
            db.rollback()
            errors.append(f"행 {row_idx}: DB 삽입 실패 (중복 가능)")
            skipped += 1

    db.commit()
    wb.close()

    logger.info(f"엑셀 임포트 완료: 전체 {total}, 생성 {created}, 건너뜀 {skipped}")
    return {"total": total, "created": created, "skipped": skipped, "errors": errors}
